import openpyxl

work_book = openpyxl.load_workbook('example.xlsx')

print(f'{type(work_book)}  \n')
print(work_book.sheetnames)

sheet = work_book.get_sheet_by_name('Sheet1')

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

print("done")
